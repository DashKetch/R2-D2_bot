import discord
from discord import app_commands
from discord.ext import commands, tasks
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timezone, time as dt_time
import os
import gspread
from google.oauth2.service_account import Credentials
from fields import *

intents = discord.Intents.default()
intents.members = True

bot = commands.Bot(command_prefix="!", intents=intents)
tree = bot.tree

_last_cleanup_date = None  # Tracks the last date the cleanup task ran to avoid multiple runs in the same day


# ── Shared helpers ────────────────────────────────────────────────────────────

def _safe_send(interaction: discord.Interaction):
    """Return the right send method depending on whether we've already deferred."""
    return interaction.followup.send if interaction.response.is_done() else interaction.response.send_message


async def _post_to_log(guild: discord.Guild, embed: discord.Embed):
    """Safely post an embed to the strike log channel."""
    log_channel = guild.get_channel(LOG_CHANNEL_ID)
    if log_channel:
        try:
            await log_channel.send(embed=embed)
        except discord.Forbidden:
            pass


async def _alert_sheets_fallback(operation: str, failures: list):
    """
    Ping CAN_VIEW_ALL_ROLE in SHEETS_ALERT_CHANNEL when one or both sheets
    could not be updated. Only called when at least one destination failed.

    failures: list of dicts with keys "target" ("Google Sheets" or "Local xlsx")
              and "error" (str).
    """
    guild = discord.utils.get(bot.guilds)
    if guild is None:
        return
    alert_channel = guild.get_channel(SHEETS_ALERT_CHANNEL)
    if alert_channel is None:
        return
    can_view_role = guild.get_role(CAN_VIEW_ALL_ROLE)
    ping  = can_view_role.mention if can_view_role else "@mods"
    names = " and ".join(f["target"] for f in failures)
    embed = discord.Embed(
        title=f"⚠️ Strike Data Could Not Be Written to {names}",
        colour=discord.Color.yellow(),
        timestamp=datetime.now(timezone.utc),
    )
    embed.add_field(name="Operation", value=operation, inline=False)
    for f in failures:
        embed.add_field(name=f"❌ {f['target']} Error", value=f"`{f['error'][:200]}`", inline=False)
    if any(f["target"] == "Google Sheets" for f in failures):
        embed.add_field(
            name="Action Required",
            value="Run `/updatestrikes` to sync the local backup to Google Sheets once it's available again.",
            inline=False,
        )
    embed.set_footer(text="Check the bot console for full error details.")
    try:
        await alert_channel.send(content=ping, embed=embed)
    except discord.Forbidden:
        pass


# ── Spreadsheet helpers ───────────────────────────────────────────────────────

HEADERS = ["Moderator", "Date command used", "Username", "Strike (1 or 2)", "Reason"]

def _get_gsheet():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds  = Credentials.from_service_account_file(GSHEET_CREDS_FILE, scopes=scopes)
    client = gspread.authorize(creds)
    sheet  = client.open_by_key(GSHEET_ID)
    ws     = sheet.get_worksheet(0)
    if ws.row_count == 0 or ws.cell(1, 1).value is None:
        ws.update("A1:E1", [HEADERS])
        ws.format("A1:E1", {
            "backgroundColor": {"red": 0.12, "green": 0.31, "blue": 0.47},
            "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
            "horizontalAlignment": "CENTER",
        })
    return ws


def _cell_style(ws, row: int, col: int, value, strike_num_col: bool = False, strike_num: int = 1):
    fill_color     = "D6E4F0" if (row % 2 == 0) else "FFFFFF"
    cell           = ws.cell(row=row, column=col, value=value)
    cell.fill      = PatternFill("solid", start_color=fill_color, end_color=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    if strike_num_col:
        cell.font = Font(name="Arial", size=10, bold=True,
                         color="C00000" if strike_num == 2 else "FF6600")
    else:
        cell.font = Font(name="Arial", size=10)
    return cell


def _init_spreadsheet():
    if os.path.exists(SPREADSHEET_PATH):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Strike Log"
    header_fill  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    col_widths = [22, 22, 22, 18, 45]
    for col_idx, (header, width) in enumerate(zip(HEADERS, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 20
    wb.save(SPREADSHEET_PATH)


def _local_append(moderator: str, username: str, strike_number: int, reason: str):
    _init_spreadsheet()
    wb = openpyxl.load_workbook(SPREADSHEET_PATH)
    ws = wb.active
    next_row = ws.max_row + 1
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=1).value is None:
            next_row = r
            break
    row_data = [moderator, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), username, strike_number, reason]
    for col_idx, value in enumerate(row_data, start=1):
        _cell_style(ws, next_row, col_idx, value,
                    strike_num_col=(col_idx == 4), strike_num=strike_number)
    wb.save(SPREADSHEET_PATH)


def _append_strike_row(moderator: str, username: str, strike_number: int, reason: str):
    import asyncio
    row_data = [
        moderator,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        username,
        strike_number,
        reason,
    ]
    failures = []
    try:
        ws = _get_gsheet()
        ws.append_row(row_data, value_input_option="USER_ENTERED")
        all_vals     = ws.col_values(1)
        new_row      = len(all_vals)
        strike_color = (
            {"red": 0.75, "green": 0.0, "blue": 0.0}
            if strike_number == 2
            else {"red": 1.0, "green": 0.4, "blue": 0.0}
        )
        ws.format(f"D{new_row}", {
            "textFormat": {"bold": True, "foregroundColor": strike_color}
        })
        print(f"[Sheets] Strike row written to Google Sheet (row {new_row})")
    except Exception as e:
        print(f"[Sheets] Google Sheets write failed: {e}")
        failures.append({"target": "Google Sheets", "error": str(e)})

    try:
        _local_append(moderator, username, strike_number, reason)
        print("[Sheets] Strike row written to local xlsx.")
    except Exception as e:
        print(f"[Sheets] Local xlsx write failed: {e}")
        failures.append({"target": "Local xlsx", "error": str(e)})

    if failures:
        asyncio.run_coroutine_threadsafe(
            _alert_sheets_fallback("Write strike row", failures), bot.loop)


def _get_strikes_for_user(target: discord.Member) -> dict:
    target_str = str(target).lower()

    def _parse_rows(rows):
        strike1_rows, strike2_rows = [], []
        for row in rows:
            if len(row) < 5:
                continue
            moderator, date, username, strike_num_raw, reason = row[:5]
            if not username:
                continue
            if str(username).lower() != target_str:
                continue
            try:
                strike_num = int(strike_num_raw)
            except (ValueError, TypeError):
                continue
            entry = {
                "date":      str(date)      if date      else "Unknown",
                "moderator": str(moderator) if moderator else "Unknown",
                "reason":    str(reason)    if reason    else "No reason given",
            }
            if strike_num == 1:
                strike1_rows.append(entry)
            elif strike_num == 2:
                strike2_rows.append(entry)
        return {
            "strike1": strike1_rows if strike1_rows else None,
            "strike2": strike2_rows if strike2_rows else None,
        }

    try:
        ws   = _get_gsheet()
        rows = ws.get_all_values()[1:]
        print(f"[Sheets] Read {len(rows)} rows from Google Sheet")
        return _parse_rows(rows)
    except Exception as e:
        print(f"[Sheets] Google Sheets read failed ({e}), reading from local xlsx.")
        _init_spreadsheet()
        wb   = openpyxl.load_workbook(SPREADSHEET_PATH, data_only=True)
        rows = [[cell for cell in row] for row in wb.active.iter_rows(min_row=2, values_only=True)]
        return _parse_rows(rows)


def _appeal_strike(target: discord.Member, appeal_num: int) -> dict:
    target_str = str(target).lower()
    try:
        ws       = _get_gsheet()
        all_rows = ws.get_all_values()
        data     = all_rows[1:]

        strike1_rows, strike2_rows = [], []
        for i, row in enumerate(data):
            if len(row) < 5 or not row[2]:
                continue
            if str(row[2]).lower() != target_str:
                continue
            try:
                snum = int(row[3])
            except (ValueError, TypeError):
                continue
            entry = {
                "sheet_row": i + 2,
                "moderator": row[0] or "Unknown",
                "date":      row[1] or "Unknown",
                "reason":    row[4] or "No reason given",
            }
            if snum == 1:
                strike1_rows.append(entry)
            elif snum == 2:
                strike2_rows.append(entry)

        if appeal_num == 1 and not strike1_rows:
            return {"success": False, "message": f"{target.mention} has no Strike 1 on record.",
                    "deleted_row": None, "demoted_row": None}
        if appeal_num == 2 and not strike2_rows:
            return {"success": False, "message": f"{target.mention} has no Strike 2 on record.",
                    "deleted_row": None, "demoted_row": None}

        deleted_row = demoted_row = None
        if appeal_num == 2:
            target_entry = strike2_rows[-1]
            deleted_row  = target_entry
            ws.delete_rows(target_entry["sheet_row"])
            print(f"[Sheets] Deleted Strike 2 row {target_entry['sheet_row']}")
        else:
            target_entry = strike1_rows[-1]
            deleted_row  = target_entry
            ws.delete_rows(target_entry["sheet_row"])
            print(f"[Sheets] Deleted Strike 1 row {target_entry['sheet_row']}")
            if strike2_rows:
                s2     = strike2_rows[-1]
                s2_row = s2["sheet_row"] - (1 if s2["sheet_row"] > target_entry["sheet_row"] else 0)
                ws.update_cell(s2_row, 4, 1)
                ws.format(f"D{s2_row}", {
                    "textFormat": {"bold": True,
                                   "foregroundColor": {"red": 1.0, "green": 0.4, "blue": 0.0}}
                })
                demoted_row = {"moderator": s2["moderator"], "date": s2["date"], "reason": s2["reason"]}
                print(f"[Sheets] Demoted Strike 2 → Strike 1 at row {s2_row}")

        cloud_result = {"success": True, "deleted_row": deleted_row, "demoted_row": demoted_row}

    except Exception as e:
        print(f"[Sheets] Google Sheets appeal failed: {e}")
        cloud_result = {"success": False, "error": str(e)}

    local_result = {"success": False, "error": "Not attempted"}
    try:
        _init_spreadsheet()
        wb       = openpyxl.load_workbook(SPREADSHEET_PATH)
        ws_local = wb.active
        l_strike1, l_strike2 = [], []

        for row_idx in range(2, ws_local.max_row + 1):
            vals = [ws_local.cell(row=row_idx, column=c).value for c in range(1, 6)]
            moderator, date, username, strike_num, reason = vals
            if username is None:
                continue
            if str(username).lower() != target_str:
                continue
            entry = {
                "row":       row_idx,
                "moderator": str(moderator) if moderator else "Unknown",
                "date":      str(date)      if date      else "Unknown",
                "reason":    str(reason)    if reason    else "No reason given",
            }
            try:
                snum = int(strike_num)
            except (ValueError, TypeError):
                continue
            if snum == 1:
                l_strike1.append(entry)
            elif snum == 2:
                l_strike2.append(entry)

        if appeal_num == 1 and not l_strike1:
            local_result = {"success": False, "error": "No Strike 1 found in local xlsx"}
        elif appeal_num == 2 and not l_strike2:
            local_result = {"success": False, "error": "No Strike 2 found in local xlsx"}
        else:
            l_deleted = l_demoted = None
            if appeal_num == 2:
                l_target  = l_strike2[-1]
                l_deleted = l_target
                ws_local.delete_rows(l_target["row"])
            else:
                l_target  = l_strike1[-1]
                l_deleted = l_target
                ws_local.delete_rows(l_target["row"])
                if l_strike2:
                    s2     = l_strike2[-1]
                    s2_row = s2["row"] - (1 if s2["row"] > l_target["row"] else 0)
                    ws_local.cell(row=s2_row, column=4, value=1)
                    _cell_style(ws_local, s2_row, 4, 1, strike_num_col=True, strike_num=1)
                    l_demoted = {"moderator": s2["moderator"], "date": s2["date"], "reason": s2["reason"]}

            for r in range(2, ws_local.max_row + 1):
                if ws_local.cell(row=r, column=1).value is None:
                    break
                for c in range(1, 6):
                    fill_color = "D6E4F0" if (r % 2 == 0) else "FFFFFF"
                    ws_local.cell(row=r, column=c).fill = PatternFill("solid", start_color=fill_color, end_color=fill_color)

            wb.save(SPREADSHEET_PATH)
            local_result = {"success": True, "deleted_row": l_deleted, "demoted_row": l_demoted}
            print("[Sheets] Appeal written to local xlsx.")

    except Exception as e:
        print(f"[Sheets] Local xlsx appeal failed: {e}")
        local_result = {"success": False, "error": str(e)}

    import asyncio
    failures = []
    if not cloud_result["success"]:
        failures.append({"target": "Google Sheets", "error": cloud_result.get("error", "Unknown error")})
    if not local_result["success"]:
        failures.append({"target": "Local xlsx", "error": local_result.get("error", "Unknown error")})
    if failures:
        asyncio.run_coroutine_threadsafe(
            _alert_sheets_fallback(f"Appeal Strike {appeal_num}", failures), bot.loop)

    if cloud_result["success"]:
        return {"success": True, "message": "Appeal processed.",
                "deleted_row": cloud_result["deleted_row"], "demoted_row": cloud_result["demoted_row"]}
    if local_result["success"]:
        return {"success": True, "message": "Appeal processed.",
                "deleted_row": local_result["deleted_row"], "demoted_row": local_result["demoted_row"]}
    return {"success": False,
            "message": f"{target.mention} — appeal could not be processed on either sheet. Check the alert channel.",
            "deleted_row": None, "demoted_row": None}


# ── Bot events ────────────────────────────────────────────────────────────────

@bot.event
async def on_ready():
    guild = discord.Object(id=GUILD_ID)
    tree.copy_global_to(guild=guild)
    await tree.sync(guild=guild)
    print(f"Logged in as {bot.user} | Slash commands synced to guild {GUILD_ID}")
    if not daily_channel_cleanup.is_running():
        daily_channel_cleanup.start()


# ── Daily channel cleanup task ─────────────────────────────────────────────────

@tasks.loop(minutes=60)
async def daily_channel_cleanup():
    global _last_cleanup_date

    now = datetime.now().astimezone()
    print(
        f"[Cleanup Debug] now={now.isoformat()} "
        f"time={now.strftime('%H:%M')} "
        f"target={CLEANUP_TIME_LOCAL}"
    )

    if now.strftime("%H:%M") != CLEANUP_TIME_LOCAL:
        return

    print("[Cleanup Debug] Time matched! Running cleanup...")

    today = now.date()
    if _last_cleanup_date == today:
        print("[Cleanup Debug] Already ran today.")
        return

    _last_cleanup_date = today

    guild = bot.get_guild(GUILD_ID)
    print(f"[Cleanup Debug] Guild found: {guild is not None}")

    if guild is None:
        print("[Cleanup] Guild not found — skipping run.")
        return

    exempt_role = guild.get_role(CLEANUP_EXEMPT_ROLE_ID)
    print(f"[Cleanup Debug] Exempt role found: {exempt_role is not None}")

    if exempt_role is None:
        print("[Cleanup] CLEANUP_EXEMPT_ROLE_ID not found — skipping run.")
        return

    print(f"[Cleanup] Starting daily cleanup across {len(CLEANUP_CHANNEL_IDS)} channel(s).")

    for channel_id in CLEANUP_CHANNEL_IDS:
        channel = guild.get_channel(channel_id)

        if channel is None:
            print(f"[Cleanup] Channel {channel_id} not found — skipping.")
            continue

        if not isinstance(channel, discord.TextChannel):
            print(f"[Cleanup] Channel {channel_id} is not a text channel — skipping.")
            continue

        def _is_not_exempt(message: discord.Message) -> bool:
            member = guild.get_member(message.author.id)

            # Purge bot/webhook/deleted users
            if member is None:
                return True

            return exempt_role not in member.roles

        try:
            deleted = await channel.purge(
                limit=None,
                check=_is_not_exempt,
                bulk=False,  # allows deleting messages older than 14 days
            )

            await channel.send(
                f"{channel.name} is an auto-cleanup channel."
            )

            print(
                f"[Cleanup] Deleted {len(deleted)} message(s) in #{channel.name}."
            )

        except discord.Forbidden:
            print(f"[Cleanup] Missing permissions to purge #{channel.name}.")

        except discord.HTTPException as e:
            print(f"[Cleanup] Failed to purge #{channel.name}: {e}")

    print("[Cleanup] Daily cleanup finished.")


# ── /strike ───────────────────────────────────────────────────────────────────

@tree.command(
    name="strike",
    description="Give a user a strike.",
    guild=discord.Object(id=GUILD_ID),
)
@app_commands.describe(username="The server member to strike", reason="Reason for the strike")
@app_commands.checks.has_permissions(moderate_members=True)
async def strike(interaction: discord.Interaction, username: discord.Member, reason: str):
    await interaction.response.defer(ephemeral=True)
    guild        = interaction.guild
    strike1_role = guild.get_role(STRIKE_1_ROLE_ID)
    strike2_role = guild.get_role(STRIKE_2_ROLE_ID)

    if not strike1_role or not strike2_role:
        await interaction.followup.send("❌ Strike roles not found. Check role IDs in `fields.py`.", ephemeral=True)
        return
    if strike2_role in username.roles:
        await interaction.followup.send(f"⚠️ {username.mention} already has **Strike 2**. No action taken.", ephemeral=True)
        return

    if strike1_role in username.roles:
        await username.add_roles(strike2_role, reason=reason)
        strike_number, role_given = 2, strike2_role
    else:
        await username.add_roles(strike1_role, reason=reason)
        strike_number, role_given = 1, strike1_role

    moderator_name = str(interaction.user)
    _append_strike_row(moderator=moderator_name, username=str(username),
                       strike_number=strike_number, reason=reason)

    colour = discord.Color.red() if strike_number == 2 else discord.Color.orange()
    embed  = discord.Embed(title=f"⚠️ Strike {strike_number} Issued", colour=colour,
                           timestamp=datetime.now(timezone.utc))
    embed.add_field(name="User",       value=username.mention,       inline=True)
    embed.add_field(name="Strike",     value=f"**{strike_number}**", inline=True)
    embed.add_field(name="Role Given", value=role_given.mention,     inline=True)
    embed.add_field(name="Reason",     value=reason,                 inline=False)
    embed.set_footer(text=f"Issued by {moderator_name}")

    log_channel = guild.get_channel(LOG_CHANNEL_ID)
    if not log_channel:
        await interaction.followup.send("❌ Log channel not found. Check `LOG_CHANNEL_ID` in `fields.py`.", ephemeral=True)
        return
    try:
        await log_channel.send(embed=embed)
    except discord.Forbidden:
        await interaction.followup.send(
            f"✅ Strike {strike_number} applied but I can't post in <#{LOG_CHANNEL_ID}> "
            f"(missing Send Messages / Embed Links).", ephemeral=True)
        return

    await interaction.followup.send(
        f"✅ **Strike {strike_number}** applied to {username.mention} and logged.", ephemeral=True)


@strike.error
async def strike_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    msg = ("❌ You don't have permission to use `/strike`."
           if isinstance(error, app_commands.MissingPermissions)
           else f"❌ Unexpected error: {error}")
    try:
        await _safe_send(interaction)(msg, ephemeral=True)
    except Exception:
        pass


# ── /viewstrikes ──────────────────────────────────────────────────────────────

@tree.command(
    name="viewstrikes",
    description="View strike history for a user. Leave username blank to check your own.",
    guild=discord.Object(id=GUILD_ID),
)
@app_commands.describe(username="The member to look up (leave blank to check yourself)")
async def viewstrikes(interaction: discord.Interaction, username: discord.Member = None):
    await interaction.response.defer(ephemeral=True)
    invoker       = interaction.user
    guild         = interaction.guild
    can_view_role = guild.get_role(CAN_VIEW_ALL_ROLE)
    target        = username or invoker
    is_self       = (target.id == invoker.id)
    has_view_role = can_view_role in invoker.roles if can_view_role else False

    if not is_self and not has_view_role:
        fail_embed = discord.Embed(title="🚫 Unauthorized /viewstrikes Attempt",
                                   colour=discord.Color.dark_red(),
                                   timestamp=datetime.now(timezone.utc))
        fail_embed.add_field(name="Attempted by", value=invoker.mention, inline=True)
        fail_embed.add_field(name="Target user",  value=target.mention,  inline=True)
        fail_embed.set_footer(text="Failed permission check")
        await _post_to_log(guild, fail_embed)
        await interaction.followup.send(
            "❌ You can only view your own strikes. "
            "You need the moderator role to view another member's strikes.", ephemeral=True)
        return

    data            = _get_strikes_for_user(target)
    strike1_entries = data["strike1"]
    strike2_entries = data["strike2"]

    colour = (discord.Color.red()    if strike2_entries else
              discord.Color.orange() if strike1_entries else
              discord.Color.green())

    embed = discord.Embed(title=f"📋 Strike Record — {target.display_name}",
                          colour=colour, timestamp=datetime.now(timezone.utc))
    embed.set_thumbnail(url=target.display_avatar.url)

    if strike1_entries:
        for i, entry in enumerate(strike1_entries, start=1):
            label = "Strike 1" if len(strike1_entries) == 1 else f"Strike 1 (#{i})"
            embed.add_field(name=f"🟠 {label}",
                            value=f"**Date:** {entry['date']}\n**Moderator:** {entry['moderator']}\n**Reason:** {entry['reason']}",
                            inline=False)
    else:
        embed.add_field(name="🟠 Strike 1", value="N/A", inline=False)

    if strike2_entries:
        for i, entry in enumerate(strike2_entries, start=1):
            label = "Strike 2" if len(strike2_entries) == 1 else f"Strike 2 (#{i})"
            embed.add_field(name=f"🔴 {label}",
                            value=f"**Date:** {entry['date']}\n**Moderator:** {entry['moderator']}\n**Reason:** {entry['reason']}",
                            inline=False)
    else:
        embed.add_field(name="🔴 Strike 2", value="N/A", inline=False)

    embed.set_footer(text=f"Requested by {str(invoker)}")
    await interaction.followup.send(embed=embed, ephemeral=True)


@viewstrikes.error
async def viewstrikes_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    try:
        await _safe_send(interaction)(f"❌ Unexpected error: {error}", ephemeral=True)
    except Exception:
        pass


# ── /strikeappeal ─────────────────────────────────────────────────────────────

@tree.command(
    name="strikeappeal",
    description="Appeal (remove) a strike from a user. Only available to moderators.",
    guild=discord.Object(id=GUILD_ID),
)
@app_commands.describe(
    username="The member whose strike is being appealed",
    strike_number="Which strike to appeal (1 or 2)",
)
async def strikeappeal(interaction: discord.Interaction,
                       username: discord.Member,
                       strike_number: int):
    await interaction.response.defer(ephemeral=True)
    invoker        = interaction.user
    guild          = interaction.guild
    can_view_role  = guild.get_role(CAN_VIEW_ALL_ROLE)
    has_permission = can_view_role in invoker.roles if can_view_role else False

    if not has_permission:
        fail_embed = discord.Embed(title="🚫 Unauthorized /strikeappeal Attempt",
                                   colour=discord.Color.dark_red(),
                                   timestamp=datetime.now(timezone.utc))
        fail_embed.add_field(name="Attempted by",    value=invoker.mention,    inline=True)
        fail_embed.add_field(name="Target user",     value=username.mention,   inline=True)
        fail_embed.add_field(name="Strike appealed", value=str(strike_number), inline=True)
        fail_embed.set_footer(text="Failed permission check")
        await _post_to_log(guild, fail_embed)
        await interaction.followup.send(
            "❌ You don't have permission to use `/strikeappeal`.", ephemeral=True)
        return

    if strike_number not in (1, 2):
        await interaction.followup.send("❌ `strike_number` must be **1** or **2**.", ephemeral=True)
        return

    result = _appeal_strike(username, strike_number)
    if not result["success"]:
        await interaction.followup.send(f"⚠️ {result['message']}", ephemeral=True)
        return

    strike1_role = guild.get_role(STRIKE_1_ROLE_ID)
    strike2_role = guild.get_role(STRIKE_2_ROLE_ID)

    if strike_number == 2:
        if strike2_role and strike2_role in username.roles:
            await username.remove_roles(strike2_role, reason=f"Strike 2 appealed by {invoker}")
    else:
        if result["demoted_row"]:
            if strike2_role and strike2_role in username.roles:
                await username.remove_roles(strike2_role,
                    reason=f"Strike 1 appealed, Strike 2 demoted to Strike 1 by {invoker}")
        else:
            if strike1_role and strike1_role in username.roles:
                await username.remove_roles(strike1_role, reason=f"Strike 1 appealed by {invoker}")

    deleted = result["deleted_row"]
    demoted = result["demoted_row"]
    embed   = discord.Embed(title=f"✅ Strike {strike_number} Appealed",
                            colour=discord.Color.green(), timestamp=datetime.now(timezone.utc))
    embed.add_field(name="User",            value=username.mention,   inline=True)
    embed.add_field(name="Strike Appealed", value=str(strike_number), inline=True)
    embed.add_field(name="Appealed by",     value=invoker.mention,    inline=True)
    if deleted:
        embed.add_field(name="Removed Entry",
                        value=f"**Date:** {deleted['date']}\n**Originally issued by:** {deleted['moderator']}\n**Reason:** {deleted['reason']}",
                        inline=False)
    if demoted:
        embed.add_field(name="⬇️ Strike 2 → Strike 1 (demoted)",
                        value=f"**Date:** {demoted['date']}\n**Originally issued by:** {demoted['moderator']}\n**Reason:** {demoted['reason']}",
                        inline=False)

    await _post_to_log(guild, embed)
    await interaction.followup.send(
        f"✅ Strike {strike_number} for {username.mention} has been appealed and the log updated.",
        ephemeral=True)


@strikeappeal.error
async def strikeappeal_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    try:
        await _safe_send(interaction)(f"❌ Unexpected error: {error}", ephemeral=True)
    except Exception:
        pass


# ── /place ────────────────────────────────────────────────────────────────────

@tree.command(
    name="place",
    description="Assign a role to a user with supporting files as evidence. Files are logged, never stored locally.",
    guild=discord.Object(id=GUILD_ID),
)
@app_commands.describe(
    username  = "The member to assign the role to",
    role      = "The role to give the user",
    file1     = "Evidence file 1 (required)",
    file2     = "Evidence file 2",
    file3     = "Evidence file 3",
    file4     = "Evidence file 4",
    file5     = "Evidence file 5",
    file6     = "Evidence file 6",
    file7     = "Evidence file 7",
    file8     = "Evidence file 8",
    file9     = "Evidence file 9",
    file10    = "Evidence file 10",
)
async def place(
    interaction : discord.Interaction,
    username    : discord.Member,
    role        : discord.Role,
    file1       : discord.Attachment,
    file2       : discord.Attachment = None,
    file3       : discord.Attachment = None,
    file4       : discord.Attachment = None,
    file5       : discord.Attachment = None,
    file6       : discord.Attachment = None,
    file7       : discord.Attachment = None,
    file8       : discord.Attachment = None,
    file9       : discord.Attachment = None,
    file10      : discord.Attachment = None,
):
    await interaction.response.defer(ephemeral=False)
    invoker        = interaction.user
    guild          = interaction.guild

    if interaction.channel_id != PLACEMENT_CMD_CHANNEL:
        await interaction.followup.send(
            f"❌ This command can only be used in <#{PLACEMENT_CMD_CHANNEL}>.",
            ephemeral=True)
        return

    can_view_role  = guild.get_role(CAN_VIEW_ALL_ROLE)
    placement_role = guild.get_role(PLACEMENT_STAFF)
    has_permission = (
        (can_view_role  and can_view_role  in invoker.roles) or
        (placement_role and placement_role in invoker.roles)
    )

    if not has_permission:
        fail_embed = discord.Embed(
            title="🚫 Unauthorized /place Attempt",
            colour=discord.Color.dark_red(),
            timestamp=datetime.now(timezone.utc),
        )
        fail_embed.add_field(name="Attempted by", value=invoker.mention,  inline=True)
        fail_embed.add_field(name="Target user",  value=username.mention, inline=True)
        fail_embed.add_field(name="Role",         value=role.mention,     inline=True)
        fail_embed.set_footer(text="Failed permission check")
        await _post_to_log(guild, fail_embed)
        await interaction.followup.send(
            "❌ You don't have permission to use `/place`.", ephemeral=True)
        return

    if role in username.roles:
        await interaction.followup.send(
            f"⚠️ {username.mention} already has {role.mention}. No action taken.")
        return

    try:
        await username.add_roles(role, reason=f"/place used by {invoker}")
    except discord.Forbidden:
        await interaction.followup.send(
            f"❌ I don't have permission to assign {role.mention}. "
            f"Make sure my role is above it in the hierarchy.", ephemeral=True)
        return

    roles_to_remove = [
        guild.get_role(rid)
        for rid in PLACEMENT_QUEUE_ROLE_IDS
        if guild.get_role(rid) and guild.get_role(rid) in username.roles
    ]
    removed_roles = []
    if roles_to_remove:
        try:
            await username.remove_roles(*roles_to_remove,
                reason=f"Placement queue cleanup by {invoker}")
            removed_roles = roles_to_remove
        except discord.Forbidden:
            pass

    attachments = [a for a in [file1, file2, file3, file4, file5,
                                file6, file7, file8, file9, file10] if a is not None]

    place_channel = guild.get_channel(PLACE_CHANNEL_ID)
    if not place_channel:
        await interaction.followup.send(
            f"✅ Role {role.mention} assigned to {username.mention}, but "
            f"`PLACE_CHANNEL_ID` is not set or the channel wasn't found.", ephemeral=True)
        return

    embed = discord.Embed(
        title="📁 Role Placement",
        colour=role.colour if role.colour.value != 0 else discord.Color.blurple(),
        timestamp=datetime.now(timezone.utc),
    )
    embed.add_field(name="User",          value=username.mention, inline=True)
    embed.add_field(name="Role Assigned", value=role.mention,     inline=True)
    embed.add_field(name="Issued by",     value=invoker.mention,  inline=True)
    embed.add_field(
        name=f"Evidence ({len(attachments)} file{'s' if len(attachments) != 1 else ''})",
        value="\n".join(f"[{a.filename}]({a.url})" for a in attachments),
        inline=False,
    )
    if removed_roles:
        embed.add_field(
            name="🗑️ Queue Roles Removed",
            value=" ".join(r.mention for r in removed_roles),
            inline=False,
        )
    embed.set_footer(text="Files are hosted on Discord's CDN — not stored locally.")

    image_exts = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
    for a in attachments:
        if os.path.splitext(a.filename)[1].lower() in image_exts:
            embed.set_image(url=a.url)
            break

    try:
        await place_channel.send(embed=embed)
    except discord.Forbidden:
        await interaction.followup.send(
            f"✅ Role assigned but I can't post in <#{PLACE_CHANNEL_ID}> "
            f"(missing Send Messages / Embed Links).", ephemeral=True)
        return

    await interaction.followup.send(
        f"✅ {role.mention} assigned to {username.mention} and logged with "
        f"{len(attachments)} file{'s' if len(attachments) != 1 else ''}.",
    )


@place.error
async def place_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    try:
        await _safe_send(interaction)(f"❌ Unexpected error: {error}", ephemeral=True)
    except Exception:
        pass


# ── /skipplace ────────────────────────────────────────────────────────────────

@tree.command(
    name="skipplace",
    description="Assign a role to a returning user who was previously placed and has rejoined.",
    guild=discord.Object(id=GUILD_ID),
)
@app_commands.describe(
    username = "The returning member to assign the role to",
    role     = "The role to give the user",
    file1    = "Evidence file (e.g. screenshot of previous placement)",
)
async def skipplace(
    interaction : discord.Interaction,
    username    : discord.Member,
    role        : discord.Role,
    file1       : discord.Attachment,
):
    await interaction.response.defer(ephemeral=False)
    invoker        = interaction.user
    guild          = interaction.guild

    if interaction.channel_id != PLACEMENT_CMD_CHANNEL:
        await interaction.followup.send(
            f"❌ This command can only be used in <#{PLACEMENT_CMD_CHANNEL}>.",
            ephemeral=True)
        return

    can_view_role  = guild.get_role(CAN_VIEW_ALL_ROLE)
    placement_role = guild.get_role(PLACEMENT_STAFF)
    has_permission = (
        (can_view_role  and can_view_role  in invoker.roles) or
        (placement_role and placement_role in invoker.roles)
    )

    if not has_permission:
        fail_embed = discord.Embed(
            title="🚫 Unauthorized /skipplace Attempt",
            colour=discord.Color.dark_red(),
            timestamp=datetime.now(timezone.utc),
        )
        fail_embed.add_field(name="Attempted by", value=invoker.mention,  inline=True)
        fail_embed.add_field(name="Target user",  value=username.mention, inline=True)
        fail_embed.add_field(name="Role",         value=role.mention,     inline=True)
        fail_embed.set_footer(text="Failed permission check")
        await _post_to_log(guild, fail_embed)
        await interaction.followup.send(
            "❌ You don't have permission to use `/skipplace`.", ephemeral=True)
        return

    if role in username.roles:
        await interaction.followup.send(
            f"⚠️ {username.mention} already has {role.mention}. No action taken.")
        return

    try:
        await username.add_roles(role, reason=f"/skipplace (returning user) used by {invoker}")
    except discord.Forbidden:
        await interaction.followup.send(
            f"❌ I don't have permission to assign {role.mention}. "
            f"Make sure my role is above it in the hierarchy.", ephemeral=True)
        return

    roles_to_remove = [
        guild.get_role(rid)
        for rid in PLACEMENT_QUEUE_ROLE_IDS
        if guild.get_role(rid) and guild.get_role(rid) in username.roles
    ]
    removed_roles = []
    if roles_to_remove:
        try:
            await username.remove_roles(*roles_to_remove,
                reason=f"Placement queue cleanup by {invoker}")
            removed_roles = roles_to_remove
        except discord.Forbidden:
            pass

    place_channel = guild.get_channel(PLACE_CHANNEL_ID)
    if not place_channel:
        await interaction.followup.send(
            f"✅ Role {role.mention} assigned to {username.mention}, but "
            f"`PLACE_CHANNEL_ID` is not set or the channel wasn't found.", ephemeral=True)
        return

    embed = discord.Embed(
        title="⏭️ Placement Skipped — Returning User",
        colour=discord.Color.gold(),
        timestamp=datetime.now(timezone.utc),
    )
    embed.add_field(name="User",          value=username.mention, inline=True)
    embed.add_field(name="Role Assigned", value=role.mention,     inline=True)
    embed.add_field(name="Issued by",     value=invoker.mention,  inline=True)
    embed.add_field(
        name="ℹ️ Reason for Skip",
        value="This user has been placed before and has since rejoined the server. "
              "Placement was skipped and their role was restored directly.",
        inline=False,
    )
    embed.add_field(name="Evidence", value=f"[{file1.filename}]({file1.url})", inline=False)
    if removed_roles:
        embed.add_field(
            name="🗑️ Queue Roles Removed",
            value=" ".join(r.mention for r in removed_roles),
            inline=False,
        )

    image_exts = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
    if os.path.splitext(file1.filename)[1].lower() in image_exts:
        embed.set_image(url=file1.url)

    embed.set_footer(text="Files are hosted on Discord's CDN — not stored locally.")

    try:
        await place_channel.send(embed=embed)
    except discord.Forbidden:
        await interaction.followup.send(
            f"✅ Role assigned but I can't post in <#{PLACE_CHANNEL_ID}> "
            f"(missing Send Messages / Embed Links).", ephemeral=True)
        return

    await interaction.followup.send(
        f"✅ {role.mention} restored to returning user {username.mention} and logged.",
    )


@skipplace.error
async def skipplace_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    try:
        await _safe_send(interaction)(f"❌ Unexpected error: {error}", ephemeral=True)
    except Exception:
        pass


# ── /updatestrikes ────────────────────────────────────────────────────────────

@tree.command(
    name="updatestrikes",
    description="Sync the local strike backup to Google Sheets. Only needed after a fallback.",
    guild=discord.Object(id=GUILD_ID),
)
async def updatestrikes(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)
    invoker        = interaction.user
    guild          = interaction.guild
    can_view_role  = guild.get_role(CAN_VIEW_ALL_ROLE)
    has_permission = can_view_role in invoker.roles if can_view_role else False

    if not has_permission:
        await interaction.followup.send(
            "❌ You don't have permission to use `/updatestrikes`.", ephemeral=True)
        return

    if not os.path.exists(SPREADSHEET_PATH):
        await interaction.followup.send(
            "ℹ️ No local backup file found — nothing to sync.", ephemeral=True)
        return

    _init_spreadsheet()
    wb         = openpyxl.load_workbook(SPREADSHEET_PATH, data_only=True)
    ws_local   = wb.active
    local_rows = []
    for row in ws_local.iter_rows(min_row=2, values_only=True):
        if row[0] is None and row[2] is None:
            continue
        local_rows.append(list(row))

    if not local_rows:
        await interaction.followup.send(
            "ℹ️ Local backup is empty — nothing to sync.", ephemeral=True)
        return

    try:
        ws_cloud = _get_gsheet()
    except Exception as e:
        await interaction.followup.send(
            f"❌ Could not connect to Google Sheets: `{e}`\nTry again once the issue is resolved.",
            ephemeral=True)
        return

    try:
        cloud_rows     = ws_cloud.get_all_values()
        last_cloud_row = len(cloud_rows)
        if last_cloud_row > 1:
            ws_cloud.delete_rows(2, last_cloud_row - 1)

        ws_cloud.append_rows(local_rows, value_input_option="USER_ENTERED")

        for i, row in enumerate(local_rows):
            sheet_row = i + 2
            try:
                strike_num = int(row[3])
            except (ValueError, TypeError):
                continue
            strike_color = (
                {"red": 0.75, "green": 0.0, "blue": 0.0}
                if strike_num == 2
                else {"red": 1.0, "green": 0.4, "blue": 0.0}
            )
            ws_cloud.format(f"D{sheet_row}", {
                "textFormat": {"bold": True, "foregroundColor": strike_color}
            })

    except Exception as e:
        await interaction.followup.send(
            f"❌ Sync failed while writing to Google Sheets: `{e}`", ephemeral=True)
        return

    alert_channel = guild.get_channel(SHEETS_ALERT_CHANNEL)
    embed = discord.Embed(
        title="✅ Google Sheets Synced from Local Backup",
        colour=discord.Color.green(),
        timestamp=datetime.now(timezone.utc),
    )
    embed.add_field(name="Rows written", value=str(len(local_rows)), inline=True)
    embed.add_field(name="Synced by",    value=invoker.mention,       inline=True)
    embed.set_footer(text="Google Sheets is now up to date with the local backup.")
    if alert_channel:
        try:
            await alert_channel.send(embed=embed)
        except discord.Forbidden:
            pass

    await interaction.followup.send(
        f"✅ Synced **{len(local_rows)}** row(s) from local backup to Google Sheets.",
        ephemeral=True,
    )


@updatestrikes.error
async def updatestrikes_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    try:
        await _safe_send(interaction)(f"❌ Unexpected error: {error}", ephemeral=True)
    except Exception:
        pass


# ── /dupcategory ──────────────────────────────────────────────────────────────

@tree.command(
    name="dupcategory",
    description="Duplicate all channels from an existing category into a new one.",
    guild=discord.Object(id=GUILD_ID),
)
@app_commands.describe(
    old_category_id    = "The ID of the category to duplicate",
    new_category_name  = "Name for the new category",
    new_category_index = "Position (index) of the new category in the channel list (0 = top)",
)
async def dupcategory(
    interaction        : discord.Interaction,
    old_category_id    : str,
    new_category_name  : str,
    new_category_index : str,
):
    await interaction.response.defer(ephemeral=True)
    invoker = interaction.user
    guild   = interaction.guild

    try:
        old_category_id = int(old_category_id)
    except ValueError:
        await interaction.followup.send(
            f"❌ `{old_category_id}` is not a valid ID. Paste the raw number, e.g. `1519168841454850089`.",
            ephemeral=True)
        return

    try:
        new_category_index = int(new_category_index)
    except ValueError:
        await interaction.followup.send(
            f"❌ `{new_category_index}` is not a valid position index. Please enter a whole number.",
            ephemeral=True)
        return

    dup_role_1 = guild.get_role(DUP_ROLE_1)
    dup_role_2 = guild.get_role(DUP_ROLE_2)
    dup_role_3 = guild.get_role(DUP_ROLE_3)
    has_permission = any(
        role and role in invoker.roles
        for role in (dup_role_1, dup_role_2, dup_role_3)
    )
    if not has_permission:
        await interaction.followup.send(
            "❌ You don't have permission to use `/dupcategory`.", ephemeral=True)
        return

    source_category = guild.get_channel(old_category_id)
    if source_category is None:
        await interaction.followup.send(
            f"❌ No channel found with ID `{old_category_id}`. "
            f"Make sure you're using the category's ID, not its name.", ephemeral=True)
        return
    if not isinstance(source_category, discord.CategoryChannel):
        await interaction.followup.send(
            f"❌ The channel with ID `{old_category_id}` is not a category.", ephemeral=True)
        return

    try:
        new_category = await guild.create_category(
            name       = new_category_name,
            position   = new_category_index,
            overwrites = source_category.overwrites,
            reason     = f"/dupcategory used by {invoker}",
        )
    except discord.Forbidden:
        await interaction.followup.send(
            "❌ I don't have permission to create categories.", ephemeral=True)
        return

    channels = sorted(source_category.channels, key=lambda c: c.position)
    created, failed = [], []

    for ch in channels:
        try:
            if isinstance(ch, discord.TextChannel):
                new_ch = await guild.create_text_channel(
                    name=ch.name, category=new_category, position=ch.position,
                    topic=ch.topic, slowmode_delay=ch.slowmode_delay, nsfw=ch.nsfw,
                    overwrites=ch.overwrites, reason=f"/dupcategory — copied from #{ch.name}")
            elif isinstance(ch, discord.VoiceChannel):
                new_ch = await guild.create_voice_channel(
                    name=ch.name, category=new_category, position=ch.position,
                    bitrate=ch.bitrate, user_limit=ch.user_limit,
                    overwrites=ch.overwrites, reason=f"/dupcategory — copied from #{ch.name}")
            elif isinstance(ch, discord.StageChannel):
                new_ch = await guild.create_stage_channel(
                    name=ch.name, category=new_category, position=ch.position,
                    overwrites=ch.overwrites, reason=f"/dupcategory — copied from #{ch.name}")
            elif isinstance(ch, discord.ForumChannel):
                new_ch = await guild.create_forum(
                    name=ch.name, category=new_category, position=ch.position,
                    topic=ch.topic, slowmode_delay=ch.slowmode_delay, nsfw=ch.nsfw,
                    overwrites=ch.overwrites, reason=f"/dupcategory — copied from #{ch.name}")
            else:
                new_ch = await guild.create_text_channel(
                    name=ch.name, category=new_category, position=ch.position,
                    overwrites=ch.overwrites, reason=f"/dupcategory — copied from #{ch.name} (fallback)")
            created.append(new_ch)
        except Exception as e:
            failed.append((ch.name, str(e)))

    summary_lines = [
        f"✅ New category **{new_category.name}** created at position {new_category_index}.",
        f"**{len(created)}/{len(channels)}** channel(s) duplicated successfully.",
    ]
    if failed:
        summary_lines.append(
            "⚠️ The following channels could not be duplicated:\n" +
            "\n".join(f"• `#{name}` — {err}" for name, err in failed)
        )
    await interaction.followup.send("\n".join(summary_lines), ephemeral=True)


@dupcategory.error
async def dupcategory_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    try:
        await _safe_send(interaction)(f"❌ Unexpected error: {error}", ephemeral=True)
    except Exception:
        pass


# ── /duprole ──────────────────────────────────────────────────────────────────

@tree.command(
    name="duprole",
    description="Duplicate an existing role's permissions into a new role with a new name, colour, and position.",
    guild=discord.Object(id=GUILD_ID),
)
@app_commands.describe(
    old_role_id = "ID of the role to copy permissions from",
    name        = "Name for the new role",
    color_code  = "Hex colour code for the new role (e.g. #FF5733 or FF5733)",
    position    = "Position of the new role in the hierarchy (1 = bottom, higher = further up)",
)
async def duprole(
    interaction : discord.Interaction,
    old_role_id : str,
    name        : str,
    color_code  : str,
    position    : str,
):
    await interaction.response.defer(ephemeral=True)
    invoker = interaction.user
    guild   = interaction.guild

    dup_role_1 = guild.get_role(DUP_ROLE_1)
    dup_role_2 = guild.get_role(DUP_ROLE_2)
    dup_role_3 = guild.get_role(DUP_ROLE_3)
    has_permission = any(
        role and role in invoker.roles
        for role in (dup_role_1, dup_role_2, dup_role_3)
    )
    if not has_permission:
        await interaction.followup.send(
            "❌ You don't have permission to use `/duprole`.", ephemeral=True)
        return

    try:
        old_role_id_int = int(old_role_id)
    except ValueError:
        await interaction.followup.send(
            f"❌ `{old_role_id}` is not a valid role ID. Paste the raw number.", ephemeral=True)
        return

    source_role = guild.get_role(old_role_id_int)
    if source_role is None:
        await interaction.followup.send(
            f"❌ No role found with ID `{old_role_id_int}`.", ephemeral=True)
        return

    try:
        position_int = int(position)
        if position_int < 1:
            raise ValueError
    except ValueError:
        await interaction.followup.send(
            f"❌ `{position}` is not a valid position. Enter a whole number ≥ 1.", ephemeral=True)
        return

    hex_str = color_code.lstrip("#").strip()
    if len(hex_str) != 6:
        await interaction.followup.send(
            f"❌ `{color_code}` is not a valid hex colour. Use a 6-digit code like `#FF5733` or `FF5733`.",
            ephemeral=True)
        return
    try:
        new_colour = discord.Color(int(hex_str, 16))
    except ValueError:
        await interaction.followup.send(
            f"❌ `{color_code}` contains invalid hex characters.", ephemeral=True)
        return

    try:
        new_role = await guild.create_role(
            name=name, permissions=source_role.permissions, colour=new_colour,
            hoist=source_role.hoist, mentionable=source_role.mentionable,
            reason=f"/duprole: copied from {source_role.name} by {invoker}",
        )
    except discord.Forbidden:
        await interaction.followup.send(
            "❌ I don't have permission to create roles. Make sure I have **Manage Roles**.",
            ephemeral=True)
        return
    except Exception as e:
        await interaction.followup.send(f"❌ Failed to create role: `{e}`", ephemeral=True)
        return

    try:
        await new_role.edit(position=position_int, reason=f"/duprole: set position to {position_int}")
    except discord.Forbidden:
        await interaction.followup.send(
            f"✅ Role {new_role.mention} created with copied permissions, but I couldn't set its "
            f"position. You can drag it manually in Server Settings → Roles.", ephemeral=True)
        return
    except Exception as e:
        await interaction.followup.send(
            f"✅ Role {new_role.mention} created, but position could not be set: `{e}`", ephemeral=True)
        return

    perm_names   = [p.replace("_", " ").title() for p, v in source_role.permissions if v]
    perm_summary = ", ".join(perm_names) if perm_names else "No permissions"
    if len(perm_summary) > 1000:
        perm_summary = perm_summary[:997] + "..."

    embed = discord.Embed(title="✅ Role Duplicated", colour=new_colour,
                          timestamp=datetime.now(timezone.utc))
    embed.add_field(name="New Role",           value=new_role.mention,                        inline=True)
    embed.add_field(name="Copied From",        value=source_role.mention,                     inline=True)
    embed.add_field(name="Position",           value=str(position_int),                       inline=True)
    embed.add_field(name="Colour",             value=f"`#{hex_str.upper()}`",                 inline=True)
    embed.add_field(name="Hoisted",            value="Yes" if source_role.hoist else "No",    inline=True)
    embed.add_field(name="Mentionable",        value="Yes" if source_role.mentionable else "No", inline=True)
    embed.add_field(name="Permissions Copied", value=perm_summary,                            inline=False)
    embed.set_footer(text=f"Created by {invoker}")
    await interaction.followup.send(embed=embed, ephemeral=True)


@duprole.error
async def duprole_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    try:
        await _safe_send(interaction)(f"❌ Unexpected error: {error}", ephemeral=True)
    except Exception:
        pass


# ── /cutchannelperms ──────────────────────────────────────────────────────────

@tree.command(
    name="cutchannelperms",
    description="Move a role's permission overwrite from one role to another in a channel, then remove the first.",
    guild=discord.Object(id=GUILD_ID),
)
@app_commands.describe(
    channel_id     = "ID of the channel to modify",
    source_role_id = "ID of the role whose permissions will be moved (and removed)",
    target_role_id = "ID of the role that will receive those permissions",
)
async def cutchannelperms(
    interaction    : discord.Interaction,
    channel_id     : str,
    source_role_id : str,
    target_role_id : str,
):
    await interaction.response.defer(ephemeral=True)
    invoker = interaction.user
    guild   = interaction.guild

    dup_role_1 = guild.get_role(DUP_ROLE_1)
    dup_role_2 = guild.get_role(DUP_ROLE_2)
    dup_role_3 = guild.get_role(DUP_ROLE_3)
    has_permission = any(
        role and role in invoker.roles
        for role in (dup_role_1, dup_role_2, dup_role_3)
    )
    if not has_permission:
        await interaction.followup.send(
            "❌ You don't have permission to use `/cutchannelperms`.", ephemeral=True)
        return

    try:
        channel_id_int = int(channel_id)
    except ValueError:
        await interaction.followup.send(f"❌ `{channel_id}` is not a valid channel ID.", ephemeral=True)
        return
    try:
        source_role_id_int = int(source_role_id)
    except ValueError:
        await interaction.followup.send(f"❌ `{source_role_id}` is not a valid role ID.", ephemeral=True)
        return
    try:
        target_role_id_int = int(target_role_id)
    except ValueError:
        await interaction.followup.send(f"❌ `{target_role_id}` is not a valid role ID.", ephemeral=True)
        return

    if source_role_id_int == target_role_id_int:
        await interaction.followup.send("❌ Source and target roles must be different.", ephemeral=True)
        return

    channel = guild.get_channel(channel_id_int)
    if channel is None:
        await interaction.followup.send(f"❌ No channel found with ID `{channel_id_int}`.", ephemeral=True)
        return
    source_role = guild.get_role(source_role_id_int)
    if source_role is None:
        await interaction.followup.send(f"❌ No role found with ID `{source_role_id_int}`.", ephemeral=True)
        return
    target_role = guild.get_role(target_role_id_int)
    if target_role is None:
        await interaction.followup.send(f"❌ No role found with ID `{target_role_id_int}`.", ephemeral=True)
        return

    source_overwrite = channel.overwrites_for(source_role)
    if source_overwrite.is_empty():
        await interaction.followup.send(
            f"⚠️ {source_role.mention} has no permission overwrite in {channel.mention}. Nothing to move.",
            ephemeral=True)
        return

    allow_perms = [p.replace("_", " ").title() for p, v in source_overwrite if v is True]
    deny_perms  = [p.replace("_", " ").title() for p, v in source_overwrite if v is False]

    try:
        await channel.set_permissions(
            target_role, overwrite=source_overwrite,
            reason=f"/cutchannelperms: moved from {source_role.name} by {invoker}")
    except discord.Forbidden:
        await interaction.followup.send(
            f"❌ I don't have permission to edit overwrites in {channel.mention}. "
            f"Make sure I have **Manage Channels** and my role is above both roles.", ephemeral=True)
        return
    except Exception as e:
        await interaction.followup.send(
            f"❌ Failed to apply permissions to {target_role.mention}: `{e}`", ephemeral=True)
        return

    try:
        await channel.set_permissions(
            source_role, overwrite=None,
            reason=f"/cutchannelperms: removed after move to {target_role.name} by {invoker}")
    except discord.Forbidden:
        await interaction.followup.send(
            f"✅ Permissions copied to {target_role.mention}, but I couldn't remove "
            f"{source_role.mention}'s overwrite. Please remove it manually.", ephemeral=True)
        return
    except Exception as e:
        await interaction.followup.send(
            f"✅ Permissions copied to {target_role.mention}, but failed to remove "
            f"{source_role.mention}'s overwrite: `{e}`", ephemeral=True)
        return

    embed = discord.Embed(title="✅ Channel Permissions Moved", colour=discord.Color.blurple(),
                          timestamp=datetime.now(timezone.utc))
    embed.add_field(name="Channel",    value=channel.mention,     inline=True)
    embed.add_field(name="Moved From", value=source_role.mention, inline=True)
    embed.add_field(name="Moved To",   value=target_role.mention, inline=True)
    if allow_perms:
        v = ", ".join(allow_perms)
        embed.add_field(name="✅ Allowed", value=v[:1021] + "..." if len(v) > 1024 else v, inline=False)
    if deny_perms:
        v = ", ".join(deny_perms)
        embed.add_field(name="❌ Denied", value=v[:1021] + "..." if len(v) > 1024 else v, inline=False)
    if not allow_perms and not deny_perms:
        embed.add_field(name="Permissions", value="Neutral (all unset)", inline=False)
    embed.set_footer(text=f"Run by {invoker}")
    await interaction.followup.send(embed=embed, ephemeral=True)


@cutchannelperms.error
async def cutchannelperms_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    try:
        await _safe_send(interaction)(f"❌ Unexpected error: {error}", ephemeral=True)
    except Exception:
        pass


# ── /roleids ──────────────────────────────────────────────────────────────────

@tree.command(
    name="roleids",
    description="Display the role ID of up to 10 mentioned roles.",
    guild=discord.Object(id=GUILD_ID),
)
@app_commands.describe(
    role1="Role 1", role2="Role 2", role3="Role 3", role4="Role 4", role5="Role 5",
    role6="Role 6", role7="Role 7", role8="Role 8", role9="Role 9", role10="Role 10",
)
async def roleids(
    interaction : discord.Interaction,
    role1       : discord.Role,
    role2       : discord.Role = None,
    role3       : discord.Role = None,
    role4       : discord.Role = None,
    role5       : discord.Role = None,
    role6       : discord.Role = None,
    role7       : discord.Role = None,
    role8       : discord.Role = None,
    role9       : discord.Role = None,
    role10      : discord.Role = None,
):
    await interaction.response.defer(ephemeral=True)
    roles = [r for r in [role1, role2, role3, role4, role5,
                          role6, role7, role8, role9, role10] if r is not None]
    lines = [f"{r.mention}` {r.id}`" for r in roles]
    embed = discord.Embed(title="🔖 Role IDs", colour=discord.Color.blurple(),
                          timestamp=datetime.now(timezone.utc))
    embed.description = "\n".join(lines)
    embed.set_footer(text=f"Requested by {interaction.user} · {len(roles)} role{'s' if len(roles) != 1 else ''}")
    await interaction.followup.send(embed=embed, ephemeral=True)


@roleids.error
async def roleids_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    try:
        await _safe_send(interaction)(f"❌ Unexpected error: {error}", ephemeral=True)
    except Exception:
        pass


# ── /channelids ───────────────────────────────────────────────────────────────

@tree.command(
    name="channelids",
    description="Display the channel ID of up to 10 mentioned channels.",
    guild=discord.Object(id=GUILD_ID),
)
@app_commands.describe(
    channel1="Channel 1", channel2="Channel 2", channel3="Channel 3",
    channel4="Channel 4", channel5="Channel 5", channel6="Channel 6",
    channel7="Channel 7", channel8="Channel 8", channel9="Channel 9", channel10="Channel 10",
)
async def channelids(
    interaction : discord.Interaction,
    channel1    : discord.abc.GuildChannel,
    channel2    : discord.abc.GuildChannel = None,
    channel3    : discord.abc.GuildChannel = None,
    channel4    : discord.abc.GuildChannel = None,
    channel5    : discord.abc.GuildChannel = None,
    channel6    : discord.abc.GuildChannel = None,
    channel7    : discord.abc.GuildChannel = None,
    channel8    : discord.abc.GuildChannel = None,
    channel9    : discord.abc.GuildChannel = None,
    channel10   : discord.abc.GuildChannel = None,
):
    await interaction.response.defer(ephemeral=True)
    channels = [c for c in [channel1, channel2, channel3, channel4, channel5,
                              channel6, channel7, channel8, channel9, channel10] if c is not None]

    def channel_icon(ch):
        if isinstance(ch, discord.VoiceChannel):   return "🔊"
        if isinstance(ch, discord.StageChannel):   return "🎙️"
        if isinstance(ch, discord.ForumChannel):   return "🗂️"
        if isinstance(ch, discord.CategoryChannel): return "📁"
        if getattr(ch, "news", False):              return "📢"
        return "💬"

    lines = [f"{channel_icon(ch)} {ch.mention}` {ch.id}`" for ch in channels]
    embed = discord.Embed(title="📋 Channel IDs", colour=discord.Color.blurple(),
                          timestamp=datetime.now(timezone.utc))
    embed.description = "\n".join(lines)
    embed.set_footer(text=f"Requested by {interaction.user} · {len(channels)} channel{'s' if len(channels) != 1 else ''}")
    await interaction.followup.send(embed=embed, ephemeral=True)


@channelids.error
async def channelids_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    try:
        await _safe_send(interaction)(f"❌ Unexpected error: {error}", ephemeral=True)
    except Exception:
        pass


# ── /cutcategoryperms ─────────────────────────────────────────────────────────

@tree.command(
    name="cutcategoryperms",
    description="Move a role's permission overwrite from one role to another in a category, then remove the first.",
    guild=discord.Object(id=GUILD_ID),
)
@app_commands.describe(
    category_id    = "ID of the category to modify",
    source_role_id = "ID of the role whose permissions will be moved (and removed)",
    target_role_id = "ID of the role that will receive those permissions",
)
async def cutcategoryperms(
    interaction    : discord.Interaction,
    category_id    : str,
    source_role_id : str,
    target_role_id : str,
):
    await interaction.response.defer(ephemeral=True)
    invoker = interaction.user
    guild   = interaction.guild

    dup_role_1 = guild.get_role(DUP_ROLE_1)
    dup_role_2 = guild.get_role(DUP_ROLE_2)
    dup_role_3 = guild.get_role(DUP_ROLE_3)
    has_permission = any(
        role and role in invoker.roles
        for role in (dup_role_1, dup_role_2, dup_role_3)
    )
    if not has_permission:
        await interaction.followup.send(
            "❌ You don't have permission to use `/cutcategoryperms`.", ephemeral=True)
        return

    try:
        category_id_int = int(category_id)
    except ValueError:
        await interaction.followup.send(f"❌ `{category_id}` is not a valid category ID.", ephemeral=True)
        return
    try:
        source_role_id_int = int(source_role_id)
    except ValueError:
        await interaction.followup.send(f"❌ `{source_role_id}` is not a valid role ID.", ephemeral=True)
        return
    try:
        target_role_id_int = int(target_role_id)
    except ValueError:
        await interaction.followup.send(f"❌ `{target_role_id}` is not a valid role ID.", ephemeral=True)
        return

    if source_role_id_int == target_role_id_int:
        await interaction.followup.send("❌ Source and target roles must be different.", ephemeral=True)
        return

    category = guild.get_channel(category_id_int)
    if category is None:
        await interaction.followup.send(f"❌ No channel found with ID `{category_id_int}`.", ephemeral=True)
        return
    if not isinstance(category, discord.CategoryChannel):
        await interaction.followup.send(
            f"❌ The channel with ID `{category_id_int}` is not a category.", ephemeral=True)
        return

    source_role = guild.get_role(source_role_id_int)
    if source_role is None:
        await interaction.followup.send(f"❌ No role found with ID `{source_role_id_int}`.", ephemeral=True)
        return
    target_role = guild.get_role(target_role_id_int)
    if target_role is None:
        await interaction.followup.send(f"❌ No role found with ID `{target_role_id_int}`.", ephemeral=True)
        return

    source_overwrite = category.overwrites_for(source_role)
    if source_overwrite.is_empty():
        await interaction.followup.send(
            f"⚠️ {source_role.mention} has no permission overwrite in {category.mention}. Nothing to move.",
            ephemeral=True)
        return

    allow_perms = [p.replace("_", " ").title() for p, v in source_overwrite if v is True]
    deny_perms  = [p.replace("_", " ").title() for p, v in source_overwrite if v is False]

    try:
        await category.set_permissions(
            target_role, overwrite=source_overwrite,
            reason=f"/cutcategoryperms: moved from {source_role.name} by {invoker}")
    except discord.Forbidden:
        await interaction.followup.send(
            f"❌ I don't have permission to edit overwrites in {category.mention}. "
            f"Make sure I have **Manage Channels** and my role is above both roles.", ephemeral=True)
        return
    except Exception as e:
        await interaction.followup.send(
            f"❌ Failed to apply permissions to {target_role.mention}: `{e}`", ephemeral=True)
        return

    try:
        await category.set_permissions(
            source_role, overwrite=None,
            reason=f"/cutcategoryperms: removed after move to {target_role.name} by {invoker}")
    except discord.Forbidden:
        await interaction.followup.send(
            f"✅ Permissions copied to {target_role.mention}, but I couldn't remove "
            f"{source_role.mention}'s overwrite. Please remove it manually.", ephemeral=True)
        return
    except Exception as e:
        await interaction.followup.send(
            f"✅ Permissions copied to {target_role.mention}, but failed to remove "
            f"{source_role.mention}'s overwrite: `{e}`", ephemeral=True)
        return

    embed = discord.Embed(title="✅ Category Permissions Moved", colour=discord.Color.blurple(),
                          timestamp=datetime.now(timezone.utc))
    embed.add_field(name="Category",   value=category.mention,    inline=True)
    embed.add_field(name="Moved From", value=source_role.mention, inline=True)
    embed.add_field(name="Moved To",   value=target_role.mention, inline=True)
    if allow_perms:
        v = ", ".join(allow_perms)
        embed.add_field(name="✅ Allowed", value=v[:1021] + "..." if len(v) > 1024 else v, inline=False)
    if deny_perms:
        v = ", ".join(deny_perms)
        embed.add_field(name="❌ Denied", value=v[:1021] + "..." if len(v) > 1024 else v, inline=False)
    if not allow_perms and not deny_perms:
        embed.add_field(name="Permissions", value="Neutral (all unset)", inline=False)
    embed.set_footer(text=f"Run by {invoker}")
    await interaction.followup.send(embed=embed, ephemeral=True)


@cutcategoryperms.error
async def cutcategoryperms_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    try:
        await _safe_send(interaction)(f"❌ Unexpected error: {error}", ephemeral=True)
    except Exception:
        pass


# ── /copyserver ───────────────────────────────────────────────────────────────

@tree.command(
    name="copyserver",
    description="Copy all roles, categories, and channels from this server into a target server.",
    guild=discord.Object(id=GUILD_ID),
)
@app_commands.describe(
    target_guild_id = "ID of the server to copy into (bot must already be in it)",
)
async def copyserver(
    interaction     : discord.Interaction,
    target_guild_id : str,
):
    await interaction.response.defer(ephemeral=True)
    invoker = interaction.user
    guild   = interaction.guild   # source

    # ── Permission check ──────────────────────────────────────────────────────
    dup_role_1 = guild.get_role(DUP_ROLE_1)
    dup_role_2 = guild.get_role(DUP_ROLE_2)
    dup_role_3 = guild.get_role(DUP_ROLE_3)
    has_permission = any(
        role and role in invoker.roles
        for role in (dup_role_1, dup_role_2, dup_role_3)
    )
    if not has_permission:
        await interaction.followup.send(
            "❌ You don't have permission to use `/copyserver`.", ephemeral=True)
        return

    # ── Validate target guild ID ──────────────────────────────────────────────
    try:
        target_guild_id_int = int(target_guild_id)
    except ValueError:
        await interaction.followup.send(
            f"❌ `{target_guild_id}` is not a valid guild ID.", ephemeral=True)
        return

    target = bot.get_guild(target_guild_id_int)
    if target is None:
        await interaction.followup.send(
            f"❌ Bot is not in a server with ID `{target_guild_id_int}`. "
            f"Invite the bot to the target server first.", ephemeral=True)
        return

    if target.id == guild.id:
        await interaction.followup.send(
            "❌ Source and target servers must be different.", ephemeral=True)
        return

    # ── Progress update helper ────────────────────────────────────────────────
    async def _progress(msg: str):
        try:
            await interaction.followup.send(msg, ephemeral=True)
        except Exception:
            pass

    await _progress(
        f"🔄 Starting copy from **{guild.name}** → **{target.name}**.\n"
        f"This may take several minutes on large servers. You\'ll get updates as each stage completes."
    )

    errors = []

    # ══════════════════════════════════════════════════════════════════════════
    # STAGE 1 — Roles
    # ══════════════════════════════════════════════════════════════════════════
    # Sort by position ascending so lower roles are created first and the
    # hierarchy ends up in the right order. Skip @everyone (it always exists).
    source_roles = sorted(
        [r for r in guild.roles if r.name != "@everyone"],
        key=lambda r: r.position,
        reverse=True  # highest position first so the stack ends up in the correct order
    )

    # role_map: source role ID → newly created target role
    role_map   = {}
    roles_ok   = 0
    roles_fail = 0

    for r in source_roles:
        try:
            new_role = await target.create_role(
                name        = r.name,
                permissions = r.permissions,
                colour      = r.colour,
                hoist       = r.hoist,
                mentionable = r.mentionable,
                reason      = f"/copyserver from {guild.name} by {invoker}",
            )
            role_map[r.id] = new_role
            roles_ok += 1
        except Exception as e:
            errors.append(f"Role `{r.name}`: {e}")
            roles_fail += 1

    # Re-apply positions.
    # source_roles is sorted ascending by position (lowest first).
    # We assign positions 1..N in that same order so the hierarchy matches.
    # discord.py expects a dict of {role: new_position}.
    try:
        position_payload = {
            role_map[r.id]: idx + 1
            for idx, r in enumerate(source_roles)
            if r.id in role_map
        }
        await target.edit_role_positions(position_payload,
                                         reason=f"/copyserver role positions from {guild.name}")
    except Exception as e:
        errors.append(f"Role position reorder: {e}")

    await _progress(f"✅ Stage 1/3 — Roles: **{roles_ok}** created, **{roles_fail}** failed.")

    # ── Overwrite translator ──────────────────────────────────────────────────
    def _translate_overwrites(source_overwrites: dict) -> dict:
        """
        Convert a channel's permission overwrites from source role/member IDs
        to the corresponding new target roles.  Member overwrites are dropped
        since members won't be the same across servers.
        """
        new_overwrites = {}
        for target_obj, overwrite in source_overwrites.items():
            if isinstance(target_obj, discord.Role):
                if target_obj.name == "@everyone":
                    # @everyone always exists; grab it from the target guild
                    new_overwrites[target.default_role] = overwrite
                elif target_obj.id in role_map:
                    new_overwrites[role_map[target_obj.id]] = overwrite
                # else: role wasn't created (failed) — skip its overwrite
            # Skip member-level overwrites — members differ between servers
        return new_overwrites

    # ══════════════════════════════════════════════════════════════════════════
    # STAGE 2 — Categories
    # ══════════════════════════════════════════════════════════════════════════
    source_categories = sorted(
        [c for c in guild.channels if isinstance(c, discord.CategoryChannel)],
        key=lambda c: c.position
    )

    # category_map: source category ID → new target category
    category_map   = {}
    cats_ok        = 0
    cats_fail      = 0

    for cat in source_categories:
        try:
            new_cat = await target.create_category(
                name       = cat.name,
                position   = cat.position,
                overwrites = _translate_overwrites(cat.overwrites),
                reason     = f"/copyserver from {guild.name} by {invoker}",
            )
            category_map[cat.id] = new_cat
            cats_ok += 1
        except Exception as e:
            errors.append(f"Category `{cat.name}`: {e}")
            cats_fail += 1

    await _progress(f"✅ Stage 2/3 — Categories: **{cats_ok}** created, **{cats_fail}** failed.")

    # ══════════════════════════════════════════════════════════════════════════
    # STAGE 3 — Channels
    # ══════════════════════════════════════════════════════════════════════════
    # Process uncategorised channels first, then categorised ones.
    source_channels = sorted(
        [c for c in guild.channels if not isinstance(c, discord.CategoryChannel)],
        key=lambda c: c.position
    )

    channels_ok   = 0
    channels_fail = 0

    for ch in source_channels:
        try:
            new_cat       = category_map.get(ch.category_id)  # None if uncategorised or category failed
            new_overwrites = _translate_overwrites(ch.overwrites)

            if isinstance(ch, discord.TextChannel):
                await target.create_text_channel(
                    name           = ch.name,
                    category       = new_cat,
                    position       = ch.position,
                    topic          = ch.topic,
                    slowmode_delay = ch.slowmode_delay,
                    nsfw           = ch.nsfw,
                    overwrites     = new_overwrites,
                    reason         = f"/copyserver from {guild.name} by {invoker}",
                )
            elif isinstance(ch, discord.VoiceChannel):
                await target.create_voice_channel(
                    name       = ch.name,
                    category   = new_cat,
                    position   = ch.position,
                    bitrate    = min(ch.bitrate, target.bitrate_limit),
                    user_limit = ch.user_limit,
                    overwrites = new_overwrites,
                    reason     = f"/copyserver from {guild.name} by {invoker}",
                )
            elif isinstance(ch, discord.StageChannel):
                await target.create_stage_channel(
                    name       = ch.name,
                    category   = new_cat,
                    position   = ch.position,
                    overwrites = new_overwrites,
                    reason     = f"/copyserver from {guild.name} by {invoker}",
                )
            elif isinstance(ch, discord.ForumChannel):
                await target.create_forum(
                    name           = ch.name,
                    category       = new_cat,
                    position       = ch.position,
                    topic          = ch.topic,
                    slowmode_delay = ch.slowmode_delay,
                    nsfw           = ch.nsfw,
                    overwrites     = new_overwrites,
                    reason         = f"/copyserver from {guild.name} by {invoker}",
                )
            else:
                # Announcement / News — fall back to text
                await target.create_text_channel(
                    name       = ch.name,
                    category   = new_cat,
                    position   = ch.position,
                    overwrites = new_overwrites,
                    reason     = f"/copyserver from {guild.name} (fallback) by {invoker}",
                )
            channels_ok += 1

        except Exception as e:
            errors.append(f"Channel `{ch.name}`: {e}")
            channels_fail += 1

    await _progress(f"✅ Stage 3/3 — Channels: **{channels_ok}** created, **{channels_fail}** failed.")

    # ── Final summary ─────────────────────────────────────────────────────────
    summary = (
        f"🏁 **Copy complete** — **{guild.name}** → **{target.name}**\n"
        f"• Roles: {roles_ok} created\n"
        f"• Categories: {cats_ok} created\n"
        f"• Channels: {channels_ok} created\n"
    )
    if errors:
        error_text = "\n".join(f"• {e}" for e in errors[:20])
        if len(errors) > 20:
            error_text += f"\n…and {len(errors) - 20} more (check console)"
        summary += f"\n⚠️ **{len(errors)} error(s):**\n{error_text}"
    else:
        summary += "\n✅ No errors."

    await _progress(summary)


@copyserver.error
async def copyserver_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    try:
        await _safe_send(interaction)(f"❌ Unexpected error: {error}", ephemeral=True)
    except Exception:
        pass


# ── Run ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    bot.run(BOT_TOKEN)